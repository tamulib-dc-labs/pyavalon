"""
Regenerate template.xlsx, the fill-in-the-blanks spreadsheet for replace_metadata.

    python scripts/make_template.py

Needs openpyxl (not a runtime dependency of pyavalon):

    pip install openpyxl

Sheets 1-2 cover replace_metadata: "Fields" documents every column the command
accepts, "Template" is the sheet to fill in and save as CSV. Sheets 3-4 do the
same for delete_supplemental_files.

Types come from the Update request body in the Avalon REST API docs:
https://samvera.atlassian.net/wiki/spaces/AVALON/pages/1957954917/REST+API

The docs badge only `title` REQUIRED. `date_issued` is marked required here as
well because the running API rejects an empty one with "Date issued field is
required." Fields the docs list but the server ignores (subject,
statement_of_responsibility) are offered but flagged, since a write to them is
reported as NOT APPLIED.
"""

import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from pyavalon.avalon.metadata import (
    DOCUMENTED_FIELDS,
    NOT_OFFERED,
    PAIRED_FIELDS,
)

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF")
REQUIRED_FILL = PatternFill("solid", fgColor="FCE4D6")
NOTE_FONT = Font(italic=True, color="7F7F7F")

# One example per documented field, keyed by the API name. The sheet itself is
# built from pyavalon's DOCUMENTED_FIELDS so the template cannot drift from the
# code, and the code is a transcription of the API docs.
EXAMPLES = {
    "title": "Adair, Thomas W. Oral History",
    "date_issued": "2000-12-05",
    "creator": "Appelt, Leslie L.",
    "alternative_title": "Adair Interview",
    "translated_title": "Entrevista con Adair",
    "uniform_title": "Oral history collection",
    "statement_of_responsibility": "by Haskell M. Monroe",
    "date_created": "2000-12-05",
    "copyright_date": "2001",
    "abstract": "An interview covering the founding of the school.",
    "note": "reformatted digital",
    "note_type": "general",
    "format": "audio/x-wav",
    "resource_type": "sound recording",
    "contributor": "Cushing Memorial Library & Archives",
    "publisher": "Texas A & M University. Libraries",
    "genre": "Oral histories",
    "subject": "Radio plays",
    "related_item_url": "https://example.org/finding-aid",
    "geographic_subject": "Texas",
    "temporal_subject": "20th century",
    "topical_subject": "Forestry",
    "bibliographic_id": "12345",
    "language": "eng",
    "terms_of_use": "No Copyright - United States",
    "table_of_contents": "Part 1 -- Part 2",
    "physical_description": "1 disc; 33 1/3 rpm",
    "other_identifier": "tfs_jek01_03b_06.wav",
    "other_identifier_type": "local",
    "comment": "Digitized from reel 4",
}

# Per-field caveats, all of them established against the live API.
CAVEATS = {
    "date_issued": "docs do not badge this REQUIRED, but Avalon rejects an empty one",
    "subject": "accepted and then ignored by Avalon - use topical_subject",
    "statement_of_responsibility": "accepted and then ignored by Avalon",
    "bibliographic_id": "cannot be cleared once set",
    "note_type": "required if note has values",
    "other_identifier_type": "required if other_identifier is used",
    "language": "controlled vocabulary - a bad value is dropped silently",
}


def required_flag(name):
    if name == "title":
        return "Y"
    if name == "date_issued":
        return "Y*"
    if name in PAIRED_FIELDS.values():
        return "Y if its partner is used"
    return "N"


def build_fields():
    """[(field, type, required, example/notes)] straight from the code."""
    rows = [("work id", "n/a - row key", "Y", "nk322d54j  (not an API field; names the object)")]
    for name, multi in DOCUMENTED_FIELDS.items():
        kind = "array<string> (repeat column)" if multi else "string"
        if name in NOT_OFFERED:
            rows.append((name, kind, "-", f"NOT A COLUMN: {NOT_OFFERED[name]}"))
            continue
        example = EXAMPLES.get(name, "")
        if name in CAVEATS:
            example = f"{example}   ({CAVEATS[name]})" if example else CAVEATS[name]
        rows.append((name, kind, required_flag(name), example))
    return rows


FIELDS = build_fields()

RULES = [
    "How replace_metadata reads this spreadsheet",
    "",
    ("1. A field is replaced ONLY if its column is present. Delete every column you do not "
     "want to touch -- a column you leave in but leave blank WIPES that field."),
    ("2. All values for a field are replaced, not merged. Repeat the column name once per "
     "value: Contributor, Contributor, Contributor gives one field three values."),
    ("3. One value per column. Values are never split on ';' or ',', so 'Lane, Daryl; Crews, "
     "David' in a single cell is ONE long name, not two."),
    "4. Quote any value containing a comma: \"Lane, Daryl\". Excel does this for you on save.",
    ("5. title and date_issued are REQUIRED by Avalon. Leaving either blank is an error and "
     "the row is skipped -- remove the column instead if you do not want to change it."),
    "6. Every other field clears when its column is blank.",
    "",
    ("Y* = the API docs badge only title as REQUIRED, but Avalon rejects an empty date_issued "
     "with \"Date issued field is required.\", so treat both as required."),
    ("subject and statement_of_responsibility are in the API docs but Avalon stores nothing "
     "when you send them - they are reported as NOT APPLIED. Use topical_subject instead of "
     "subject."),
    ("Column names are the field names from the API docs, so what you type here is what you "
     "read there. Title Case with spaces works too - 'date_issued' and 'Date Issued' are the "
     "same column - so a sheet exported from Avalon can be fed back in unchanged."),
    ("work id is not a metadata field - it names which object to update and goes in the URL."),
    ("Avalon can accept a value and store nothing -- Note Type and Language are controlled "
     "vocabularies, and a bad value is dropped silently. Every write is read back, and "
     "anything that did not stick is reported as NOT APPLIED."),
    "",
    "Save Sheet 2 as CSV, then run:",
    "    pyavalon replace_metadata -c yourfile.csv -i pre --dry_run",
    "Check the report, then drop --dry_run to apply.",
]

# Columns pre-laid-out on the template sheet. Repeats give room for several
# values; delete any column you are not changing.
TEMPLATE_COLUMNS = [
    "work id", "title", "date_issued",
    "creator", "creator", "creator",
    "contributor", "contributor", "contributor",
    "publisher", "publisher",
    "genre", "genre", "genre",
    "abstract",
]


# --- delete_supplemental_files -------------------------------------------
# type label, what it deletes, notes
DELETE_TYPES = [
    ("caption", "supplementals stored as \"caption\"", "also accepts 'captions'"),
    ("transcript", "supplementals stored as \"transcript\"",
     "also accepts 'transcripts'. A caption flagged treat_as_transcript stays a caption"),
    ("generic", "supplementals stored as \"generic\"",
     "also accepts 'generics'. PDFs live here, along with every other non-caption attachment"),
    ("pdf", "REFUSED - not an Avalon type",
     "PDFs are stored as 'generic'; use that if you mean every attachment of that kind"),
]

DELETE_RULES = [
    "How delete_supplemental_files reads this spreadsheet",
    "",
    ("1. One row per work per type. The id is the WORK id, not a master file id -- supplemental "
     "files hang off master files, so the work is read first and every master file on it is swept."),
    ("2. Every supplemental of that type on that work is deleted. There is no way to pick out "
     "individual files; use the Avalon UI for that."),
    ("3. THIS CANNOT BE UNDONE. Avalon returns no response body on a delete and the file is gone. "
     "Always run with --dry_run first and read the report."),
    ("4. A work where nothing matched still gets a report row, so a typo'd work id does not look "
     "like a clean sweep."),
    "",
    "Save Sheet 4 as CSV, then run:",
    "    pyavalon delete_supplemental_files -c yourfile.csv -i pre --dry_run",
    "Check the report, then drop --dry_run to delete.",
]

DELETE_TEMPLATE_COLUMNS = ["work id", "type"]


def build():
    book = Workbook()

    fields = book.active
    fields.title = "Fields"
    fields.append(["field", "type", "required Y/N", "example"])
    for cell in fields[1]:
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
    for row in FIELDS:
        fields.append(list(row))
        if row[2] != "N":
            for cell in fields[fields.max_row]:
                cell.fill = REQUIRED_FILL
    for column, width in zip("ABCD", (30, 24, 32, 48)):
        fields.column_dimensions[column].width = width
    fields.freeze_panes = "A2"

    fields.append([])
    for line in RULES:
        fields.append([line])
        cell = fields.cell(row=fields.max_row, column=1)
        cell.font = Font(bold=True) if line.startswith("How ") else NOTE_FONT
        cell.alignment = Alignment(wrap_text=False)

    template = book.create_sheet("Template")
    template.append(TEMPLATE_COLUMNS)
    for cell in template[1]:
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    template.append([
        "nk322d54j", "Adair, Thomas W. Oral History", "2000-12-05",
        "Appelt, Leslie L.", "", "",
        "Monroe, Haskell M.", "Cushing Memorial Library & Archives", "",
        "Texas A & M University. Libraries", "",
        "Interview", "Oral histories", "",
        "",
    ])
    for cell in template[2]:
        cell.font = NOTE_FONT
    template.append([])
    template.cell(row=4, column=1, value="^ row 2 is an example - delete it before running").font = NOTE_FONT
    for index in range(1, len(TEMPLATE_COLUMNS) + 1):
        template.column_dimensions[get_column_letter(index)].width = 26
    template.freeze_panes = "B2"

    types = book.create_sheet("Delete Types")
    types.append(["type", "deletes", "notes"])
    for cell in types[1]:
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
    for row in DELETE_TYPES:
        types.append(list(row))
        if row[0] == "pdf":
            for cell in types[types.max_row]:
                cell.fill = REQUIRED_FILL
    for column, width in zip("ABC", (18, 40, 74)):
        types.column_dimensions[column].width = width
    types.freeze_panes = "A2"
    types.append([])
    for line in DELETE_RULES:
        types.append([line])
        cell = types.cell(row=types.max_row, column=1)
        cell.font = Font(bold=True) if line.startswith("How ") else NOTE_FONT

    deletions = book.create_sheet("Delete Template")
    deletions.append(DELETE_TEMPLATE_COLUMNS)
    for cell in deletions[1]:
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    deletions.append(["vh53ww06b", "transcript"])
    for cell in deletions[2]:
        cell.font = NOTE_FONT
    deletions.append([])
    deletions.cell(row=4, column=1,
                   value="^ row 2 is an example - delete it before running").font = NOTE_FONT
    for index in range(1, len(DELETE_TEMPLATE_COLUMNS) + 1):
        deletions.column_dimensions[get_column_letter(index)].width = 26
    deletions.freeze_panes = "A2"

    book.save("template.xlsx")
    print(f"wrote template.xlsx: {len(FIELDS)} metadata fields, "
          f"{len(DELETE_TYPES)} deletion types, {len(book.sheetnames)} sheets")


if __name__ == "__main__":
    build()
